import pandas as pd
import json
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Load your matching results
with open('matched_results.json', 'r') as f:
    data = json.load(f)

# Create a matrix of candidates vs. jobs
candidates = set()
jobs = set()
match_scores = {}

for resume_path, job_matches in data.get('job_matches', {}).items():
    for job in job_matches:
        candidate_name = job.get('candidate_name', 'Unknown')
        job_title = job.get('job_title', 'Unknown')
        job_company = job.get('job_company', 'Unknown')
        match_score = job.get('score', 0)

        candidates.add(candidate_name)
        jobs.add(f"{job_title} ({job_company})")

        match_scores[(candidate_name, f"{job_title} ({job_company})")] = match_score

# Create a DataFrame
df = pd.DataFrame(
    index=sorted(candidates),
    columns=sorted(jobs),
    dtype=float
)

# Fill in the match scores
for (candidate, job), score in match_scores.items():
    df.at[candidate, job] = score * 100  # Convert to percentage

# Replace NaN with 0
df = df.fillna(0)

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Match Matrix"

# Add title
ws.merge_cells('A1:E1')
ws['A1'] = "Resume-Job Match Matrix"
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

# Add header row for date and time
from datetime import datetime

ws.merge_cells('A2:E2')
ws['A2'] = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws['A2'].font = Font(italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Add dataframe to worksheet starting at row 4
for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 4):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # Format the header row and index column
        if r_idx == 4 or c_idx == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Format the score cells
        if r_idx > 4 and c_idx > 1:
            cell.number_format = '0.0'
            # We'll apply conditional formatting to these cells later

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)

    for cell in column:
        if cell.value:
            cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)

    adjusted_width = max_length + 2
    ws.column_dimensions[column_letter].width = min(adjusted_width, 30)

# Add conditional formatting (color scale)
first_data_cell = ws.cell(row=5, column=2)
last_data_cell = ws.cell(row=ws.max_row, column=ws.max_column)
color_scale_range = f"{first_data_cell.coordinate}:{last_data_cell.coordinate}"

ws.conditional_formatting.add(
    color_scale_range,
    ColorScaleRule(
        start_type='num', start_value=0, start_color='FFFF0000',  # Red for low scores
        mid_type='num', mid_value=50, mid_color='FFFFFF00',  # Yellow for middle scores
        end_type='num', end_value=100, end_color='FF00FF00'  # Green for high scores
    )
)

# Add a legend
ws['A' + str(ws.max_row + 2)] = "Color Legend:"
ws['A' + str(ws.max_row)].font = Font(bold=True)

ws['A' + str(ws.max_row + 3)] = "0-50%: Poor Match"
ws['B' + str(ws.max_row)] = ""
ws['B' + str(ws.max_row)].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

ws['A' + str(ws.max_row + 4)] = "50-75%: Average Match"
ws['B' + str(ws.max_row)] = ""
ws['B' + str(ws.max_row)].fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

ws['A' + str(ws.max_row + 5)] = "75-100%: Strong Match"
ws['B' + str(ws.max_row)] = ""
ws['B' + str(ws.max_row)].fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")

# Add a best matches summary sheet
best_matches_ws = wb.create_sheet(title="Best Matches")
best_matches_ws['A1'] = "Top 3 Candidates for Each Job Position"
best_matches_ws['A1'].font = Font(size=16, bold=True)

# Create a best matches summary
row = 3
for job in sorted(jobs):
    # Get top 3 candidates for this job
    job_scores = [(candidate, df.at[candidate, job]) for candidate in df.index]
    top_candidates = sorted(job_scores, key=lambda x: x[1], reverse=True)[:3]

    best_matches_ws['A' + str(row)] = f"Job: {job}"
    best_matches_ws['A' + str(row)].font = Font(bold=True)
    row += 1

    best_matches_ws['A' + str(row)] = "Candidate"
    best_matches_ws['B' + str(row)] = "Match Score (%)"
    best_matches_ws['A' + str(row)].font = Font(bold=True)
    best_matches_ws['B' + str(row)].font = Font(bold=True)
    row += 1

    for candidate, score in top_candidates:
        best_matches_ws['A' + str(row)] = candidate
        best_matches_ws['B' + str(row)] = score
        best_matches_ws['B' + str(row)].number_format = '0.0'
        row += 1

    row += 2  # Add space between job positions

# Auto-adjust column widths in the best matches sheet
for column in best_matches_ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)

    for cell in column:
        if cell.value:
            cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)

    adjusted_width = max_length + 2
    best_matches_ws.column_dimensions[column_letter].width = min(adjusted_width, 30)

# Save the Excel file
wb.save("resume_job_matches.xlsx")
print("Excel report generated: resume_job_matches.xlsx")